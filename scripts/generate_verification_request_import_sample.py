from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "location_name",
    "provider_name",
    "appointment_date",
    "appointment_time",
    "priority",
    "form_type",
    "patient_full_name",
    "patient_dob",
    "patient_identifier",
    "patient_zip",
    "pms_id",
    "is_pre_registered",
    "payer_name",
    "member_id",
    "group_number",
    "subscriber_name",
    "subscriber_dob",
    "plan_priority",
    "notes",
]

EXAMPLE_ROW = [
    "New York",
    "Dr. Emma Carter",
    "2026-05-21",
    "09:30",
    "normal",
    "full_form",
    "Liam Bennett",
    "1992-08-14",
    "U63292952",
    "10001",
    "PMS-DEMO-1001",
    "yes",
    "Delta Dental of Kentucky",
    "U63292952",
    "DD-4201",
    "Liam Bennett",
    "1992-08-14",
    "primary",
    "Please verify active coverage and annual maximum before the visit.",
]


def build_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Verification Requests"

    sheet.append(HEADERS)
    sheet.append(EXAMPLE_ROW)

    header_fill = PatternFill("solid", fgColor="F59E0B")
    header_font = Font(bold=True, color="FFFFFF")

    for idx, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[get_column_letter(idx)].width = max(18, len(header) + 3)

    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "Verification Request Import"
    instructions["A1"].font = Font(bold=True, size=14)
    notes = [
        "Fill one verification request per row in the Verification Requests sheet.",
        "Use dates in YYYY-MM-DD format and time in HH:MM (24-hour) or HH:MM AM/PM format.",
        "location_name and provider_name must match records inside the selected clinic scope.",
        "If the clinic has purchased managed verification service, imported requests will appear in the Admin queue automatically.",
        "If no managed verification service enrollment is active, imported requests remain clinic self-service.",
        "patient_identifier can be member ID or another patient identifier used by the clinic.",
        "Leave pms_id blank if the clinic does not use the PMS modules.",
    ]
    for row_number, note in enumerate(notes, start=3):
        instructions[f"A{row_number}"] = note

    instructions.column_dimensions["A"].width = 120

    return workbook


def main() -> None:
    output_path = Path(r"C:\xampp\htdocs\prodentalemr\verification-request-import-sample.xlsx")
    workbook = build_workbook()
    workbook.save(output_path)


if __name__ == "__main__":
    main()
